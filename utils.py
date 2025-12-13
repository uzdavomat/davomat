import hashlib
import qrcode
import os
from datetime import datetime, timedelta, date, time
from config import SECRET_KEY, HOURLY_RATE, LUNCH_START, LUNCH_END, LATE_TIME_LIMIT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# --- QR KOD MANTIQI (O'zgarishsiz) ---
def get_daily_token(action_type):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    raw_string = f"{action_type}|{timestamp}|{SECRET_KEY}"
    token_hash = hashlib.sha256(raw_string.encode()).hexdigest()[:10]
    return f"{action_type}_{timestamp[:8]}_{token_hash}"


def verify_token(token_string):
    try:
        parts = token_string.split('_')
        if len(parts) != 3: return False, None
        action_type, date_prefix, token_hash = parts

        if action_type not in ["in", "out"]:
            return False, "Noto'g'ri amal turi!"

        return True, action_type
    except:
        return False, "Xatolik"


def generate_qr_image(bot_username, action_type):
    token = get_daily_token(action_type)
    qr_data = f"https://t.me/{bot_username}?start={token}"

    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')
    path = f"qr_{action_type}.png"
    img.save(path)
    return path


# --- VAQT HISOBLASH VA MAOSH MANTIQI ---

def is_sunday(date_str):
    """Berilgan sana yakshanba ekanligini tekshiradi."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.weekday() == 6  # Yakshanba (0=Dushanba, 6=Yakshanba)
    except:
        return False


def get_time_object(time_str):
    """Vaqt satrini datetime.time obyektiga aylantiradi."""
    return datetime.strptime(time_str, "%H:%M:%S").time()


def calculate_work_duration(check_in_str, check_out_str, date_str, late_time_limit_str=LATE_TIME_LIMIT):
    if not check_in_str:
        return "Yo'q", False, 0.0  # Duration, Is_late, Worked_seconds

    check_in_dt = datetime.strptime(check_in_str, "%H:%M:%S")
    late_limit_time = get_time_object(late_time_limit_str)

    is_late = check_in_dt.time() > late_limit_time

    if not check_out_str:
        return "N/A", is_late, 0.0  # Ketdi qilmagan

    check_out_dt = datetime.strptime(check_out_str, "%H:%M:%S")

    # Ishlash vaqti (1900-01-01 sana bilan)
    start = check_in_dt
    end = check_out_dt

    if end < start:
        end += timedelta(days=1)  # Yarim tundan keyin chiqsa, bir kun qo'shish

    total_duration = end - start
    lunch_duration = timedelta()

    # 2. Tushlik vaqtini chegirib tashlash mantiqi (XATO TUZATILGAN QISM)
    lunch_start_time = get_time_object(LUNCH_START)
    lunch_end_time = get_time_object(LUNCH_END)

    # Davomat yozuvi sanasini aniqlash
    try:
        current_day = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        current_day = date.today()

    # Tushlik oralig'ini davomat yozuvi sanasi bilan to'liq datetime obyektiga aylantirish
    lunch_start_dt = datetime.combine(current_day, lunch_start_time)
    lunch_end_dt = datetime.combine(current_day, lunch_end_time)

    # start va end ni haqiqiy sana bilan birlashtirish
    start_with_date = datetime.combine(current_day, start.time())
    end_with_date = datetime.combine(current_day, end.time())

    # Agar chiqish vaqti kirish vaqtidan oldin bo'lsa (yarim tundan keyin), bir kun qo'shamiz
    if end_with_date < start_with_date:
        end_with_date += timedelta(days=1)

    # Tushlik oralig'ining ish vaqti bilan kesishishini hisoblash
    overlap_start = max(start_with_date, lunch_start_dt)
    overlap_end = min(end_with_date, lunch_end_dt)

    if overlap_end > overlap_start:
        lunch_duration = overlap_end - overlap_start

    # Sof ishlagan vaqt
    net_duration = total_duration - lunch_duration
    total_seconds = net_duration.total_seconds()

    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)

    duration_str = f"{hours}s {minutes}m"

    return duration_str, is_late, total_seconds


def calculate_salary_per_day(date_str, worked_seconds):
    """Kunlik maoshni va daqiqalik maoshni hisoblaydi."""
    if is_sunday(date_str) or worked_seconds == 0:
        return 0.0, 0.0  # Kunlik maosh, Minutlik maosh

    worked_hours = worked_seconds / 3600

    # Maksimal to'lanadigan soatlar (Masalan, 9:00 dan 19:00 gacha - 1 soat tushlik = 9 soat)
    max_paid_hours = 9.0

    paid_hours = min(worked_hours, max_paid_hours)

    salary_per_day = paid_hours * HOURLY_RATE

    # Minutlik maoshni hisoblash
    paid_minutes = paid_hours * 60
    salary_per_minute = salary_per_day / paid_minutes if paid_minutes > 0 else 0.0

    return round(salary_per_day, 0), round(salary_per_minute, 2)


# --- EXCEL EXPORT FUNKSIYASI (O'zgarishsiz) ---

def export_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat Hisoboti"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # YANGI USTUNLAR: Kunlik va Minutlik Maosh
    headers = ["F.I.SH", "Sana", "Kelgan Vaqt", "Ketgan Vaqt", "Ishlagan Soat (Net)", "Kechikish",
               "Kunlik Maosh (so'm)", "Minutlik Maosh (so'm)"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_num)].width = 15

    total_monthly_salary = {}

    for row_num, row_data in enumerate(data, 2):
        name, date_str, check_in, check_out = row_data

        duration_str, is_late, worked_seconds = calculate_work_duration(check_in, check_out, date_str)

        salary_per_day, salary_per_minute = calculate_salary_per_day(date_str, worked_seconds)

        if name not in total_monthly_salary:
            total_monthly_salary[name] = 0
        total_monthly_salary[name] += salary_per_day

        excel_row = [
            name,
            date_str,
            check_in if check_in else "Yo'q",
            check_out if check_out else "Yo'q",
            f"{worked_seconds / 3600:.2f}" if worked_seconds > 0 else "0.00",
            "✅ Ha" if is_late else "❌ Yo'q",
            f"{salary_per_day:,.0f}",
            f"{salary_per_minute:.2f}"
        ]

        ws.append(excel_row)

        for col_num in range(1, len(excel_row) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num == 6 and is_late:
                cell.font = Font(bold=True, color="FF0000")
            elif col_num == 6 and not is_late:
                cell.font = Font(color="008000")
            if col_num == 7 and salary_per_day == 0 and not is_sunday(date_str):
                cell.font = Font(color="808080")

                # --- Oylik umumiy hisobotni qo'shish (Eng pastda) ---
    ws.append([])
    ws.append(["UMUMIY OYLIK MAOSH HISOBOTI:", "", "", "", "", "", "", ""])

    current_row = ws.max_row

    for name, salary in total_monthly_salary.items():
        current_row += 1
        ws.cell(row=current_row, column=1, value=name).font = bold_font
        ws.cell(row=current_row, column=7, value=f"Jami Maosh: {salary:,.0f} so'm").font = bold_font

    filename = f"Davomat_Hisoboti_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    return filename